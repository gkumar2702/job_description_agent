"""
Question Bank component for managing, deduplicating, scoring, and exporting interview questions.
"""

import os
import csv
import json
import asyncio
from typing import List, Dict, Any, Optional
from datetime import datetime
from collections import defaultdict
from rapidfuzz import fuzz
import aiofiles  # type: ignore
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from ..utils.config import Config
from ..utils.logger import get_logger
from ..utils.embeddings import compute_similarity
from ..utils.constants import SIMILARITY_THRESHOLD
from ..utils.decorators import log_time, log_time_async
from ..utils.schemas import Question
from .jd_parser import JobDescription
from .scoring_strategies import ScoringStrategy, HeuristicScorer
from .pdf_exporter import PDFExporter

logger = get_logger(__name__)


class QuestionBank:
    """Manages and exports interview questions."""
    
    def __init__(self, config: Optional[Config] = None, scorer: Optional[ScoringStrategy] = None):
        """
        Initialize the question bank.
        
        Args:
            config: Configuration object
            scorer: Scoring strategy to use (defaults to HeuristicScorer)
        """
        self.questions: List[Question] = []
        if config is None:
            config = Config()
        self.export_dir = config.get_export_dir()
        os.makedirs(self.export_dir, exist_ok=True)
        
        # Use provided scorer or default to HeuristicScorer
        self.scorer = scorer if scorer is not None else HeuristicScorer()
        self.pdf_exporter = PDFExporter(self.export_dir)
    
    def add_questions(self, questions: List[Dict[str, Any]]) -> None:
        """
        Add questions to the question bank.
        
        Args:
            questions: List of question dictionaries
        """
        # Convert dict questions to Question objects
        question_objects = []
        for q_dict in questions:
            try:
                question_obj = Question(**q_dict)
                question_objects.append(question_obj)
            except Exception as e:
                logger.warning(f"Failed to create Question object: {e}")
                continue
        
        self.questions.extend(question_objects)
        logger.info("questions_added", count=len(question_objects))
    
    @log_time("dedup_done")
    def deduplicate_questions(self) -> List[Question]:
        """
        Remove duplicate questions based on similarity.
        
        Returns:
            List[Question]: Deduplicated questions
        """
        if not self.questions:
            return []
        
        before_count = len(self.questions)
        
        # Group questions by difficulty and category
        grouped_questions = defaultdict(list)
        for question in self.questions:
            key = (question.difficulty, question.category)
            grouped_questions[key].append(question)
        
        deduplicated = []
        
        for (difficulty, category), questions in grouped_questions.items():
            # Remove exact duplicates first
            unique_questions = self._remove_exact_duplicates(questions)
            
            # Remove similar questions
            final_questions = self._remove_similar_questions(unique_questions)
            
            deduplicated.extend(final_questions)
        
        after_count = len(deduplicated)
        logger.info("dedup_done", before=before_count, after=after_count)
        
        self.questions = deduplicated
        return deduplicated
    
    def _remove_exact_duplicates(self, questions: List[Question]) -> List[Question]:
        """
        Remove exact duplicate questions.
        
        Args:
            questions: List of questions
            
        Returns:
            List[Question]: Questions without exact duplicates
        """
        seen_questions = set()
        unique_questions = []
        
        for question in questions:
            # Create a normalized version of the question for comparison
            normalized = self._normalize_question(question.question)
            
            if normalized not in seen_questions:
                seen_questions.add(normalized)
                unique_questions.append(question)
        
        return unique_questions
    
    def _normalize_question(self, question: str) -> str:
        """
        Normalize a question for comparison.
        
        Args:
            question: Question text
            
        Returns:
            str: Normalized question
        """
        # Convert to lowercase
        normalized = question.lower()
        
        # Remove punctuation
        import re
        normalized = re.sub(r'[^\w\s]', '', normalized)
        
        # Remove extra whitespace
        normalized = ' '.join(normalized.split())
        
        return normalized
    
    def _remove_similar_questions(self, questions: List[Question], 
                                similarity_threshold: int = SIMILARITY_THRESHOLD) -> List[Question]:
        """
        Remove similar questions based on content similarity using rapidfuzz.
        
        Args:
            questions: List of questions
            similarity_threshold: Threshold for similarity (0-100)
            
        Returns:
            List[Question]: Questions without similar duplicates
        """
        if len(questions) <= 1:
            return questions
        
        # Use rapidfuzz token_set_ratio for similarity checking
        unique_questions = []
        
        for i, question in enumerate(questions):
            is_similar = False
            
            for j, existing_question in enumerate(unique_questions):
                similarity = self._calculate_similarity(
                    question.question,
                    existing_question.question
                )
                
                if similarity >= similarity_threshold:
                    is_similar = True
                    break
            
            if not is_similar:
                unique_questions.append(question)
        
        return unique_questions
    
    def _calculate_similarity(self, question1: str, question2: str) -> int:
        """
        Calculate similarity between two questions using rapidfuzz token_set_ratio.
        
        Args:
            question1: First question
            question2: Second question
            
        Returns:
            int: Similarity score (0-100)
        """
        if not question1 or not question2:
            return 0
        
        # Use rapidfuzz token_set_ratio for better similarity detection
        # This handles paraphrasing, word order changes, and partial matches
        return int(round(fuzz.token_set_ratio(question1, question2)))
    
    @log_time("scoring_done")
    def score_questions(self, jd: JobDescription) -> List[Dict[str, Any]]:
        """
        Score questions based on relevance to job description.
        
        Args:
            jd: Job description object
            
        Returns:
            List[Dict[str, Any]]: Questions with relevance scores
        """
        scored_questions = []
        
        for question in self.questions:
            score = self.scorer.score(question, jd)
            scored_question = question.model_dump()
            scored_question['relevance_score'] = score
            scored_question['score'] = score
            scored_questions.append(scored_question)
        
        # Sort by relevance score (highest first)
        scored_questions.sort(key=lambda x: x.get('relevance_score', 0), reverse=True)
        
        logger.info("scoring_done", count=len(scored_questions))
        return scored_questions
    
    async def export_questions(self, jd: JobDescription, 
                             questions: List[Dict[str, Any]], 
                             formats: List[str] = None) -> Dict[str, str]:
        """
        Export questions as a single PDF file only.
        """
        return self._export_pdf_only(jd, questions)
    
    def export_questions_sync(self, jd: JobDescription, 
                            questions: List[Dict[str, Any]], 
                            formats: List[str] = None) -> Dict[str, str]:
        """Synchronous PDF-only export for backward compatibility."""
        return self._export_pdf_only(jd, questions)
    
    async def export_questions_async(self, jd: JobDescription, 
                                   questions: List[Dict[str, Any]], 
                                   formats: List[str] = None) -> Dict[str, str]:
        """Asynchronous wrapper that still performs PDF-only export."""
        return self._export_pdf_only(jd, questions)
    
    def _export_pdf_only(self, jd: JobDescription, questions: List[Dict[str, Any]]) -> Dict[str, str]:
        """Generate a single PDF export and return its path under the 'pdf' key."""
        try:
            metadata = {
                'total_questions': len(questions),
                'questions_by_difficulty': self._count_by_difficulty_from_list(questions),
                'average_relevance_score': self._avg_relevance_from_list(questions),
                'generated_at': datetime.now().isoformat(),
            }
            pdf_path = self.pdf_exporter.export_questions_to_pdf(jd, questions, metadata)
            logger.info(f"Exported questions in pdf format: {pdf_path}")
            return {'pdf': pdf_path}
        except Exception as e:
            logger.error(f"Error exporting PDF: {e}")
            return {}

    def _count_by_difficulty_from_list(self, questions: List[Dict[str, Any]]) -> Dict[str, int]:
        counts: Dict[str, int] = defaultdict(int)
        for q in questions:
            counts[q.get('difficulty', 'unknown')] += 1
        return dict(counts)

    def _avg_relevance_from_list(self, questions: List[Dict[str, Any]]) -> float:
        scores = [q.get('relevance_score') for q in questions if isinstance(q.get('relevance_score'), (int, float))]
        return sum(scores) / len(scores) if scores else 0.0
    
    def _export_markdown(self, questions: List[Dict[str, Any]], 
                        jd: JobDescription, filename_base: str) -> str:
        """
        Export questions in Markdown format.
        
        Args:
            questions: List of questions
            jd: Job description object
            filename_base: Base filename
            
        Returns:
            str: Path to exported file
        """
        file_path = os.path.join(self.export_dir, f"{filename_base}.md")
        
        with open(file_path, 'w', encoding='utf-8') as f:
            # Header
            f.write(f"# Interview Questions for {jd.company} - {jd.role}\n\n")
            f.write(f"**Generated on:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"**Total Questions:** {len(questions)}\n\n")
            
            # Job description summary
            f.write("## Job Description Summary\n\n")
            f.write(f"- **Company:** {jd.company}\n")
            f.write(f"- **Role:** {jd.role}\n")
            f.write(f"- **Location:** {jd.location}\n")
            f.write(f"- **Experience Required:** {jd.experience_years} years\n")
            f.write(f"- **Key Skills:** {', '.join(jd.skills[:10])}\n\n")
            
            # Group questions by difficulty
            questions_by_difficulty = defaultdict(list)
            for question in questions:
                difficulty = question.get('difficulty', 'unknown')
                questions_by_difficulty[difficulty].append(question)
            
            # Export questions by difficulty
            for difficulty in ['easy', 'medium', 'hard']:
                if difficulty in questions_by_difficulty:
                    f.write(f"## {difficulty.title()} Questions\n\n")
                    
                    for i, question in enumerate(questions_by_difficulty[difficulty], 1):
                        f.write(f"### {i}. {question.get('question', '')}\n\n")
                        f.write(f"**Answer:** {question.get('answer', '')}\n\n")
                        f.write(f"**Category:** {question.get('category', 'Technical')}\n")
                        f.write(f"**Skills:** {', '.join(question.get('skills', []))}\n")
                        f.write(f"**Source:** {question.get('source', 'Generated')}\n")
                        if 'relevance_score' in question:
                            f.write(f"**Relevance Score:** {question['relevance_score']:.2f}\n")
                        f.write("\n---\n\n")
        
        return file_path
    
    async def _export_markdown_async(self, questions: List[Dict[str, Any]], 
                                   jd: JobDescription, filename_base: str) -> str:
        """
        Export questions in Markdown format asynchronously.
        
        Args:
            questions: List of questions
            jd: Job description object
            filename_base: Base filename
            
        Returns:
            str: Path to exported file
        """
        file_path = os.path.join(self.export_dir, f"{filename_base}.md")
        
        content = []
        
        # Header
        content.append(f"# Interview Questions for {jd.company} - {jd.role}\n\n")
        content.append(f"**Generated on:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        content.append(f"**Total Questions:** {len(questions)}\n\n")
        
        # Job description summary
        content.append("## Job Description Summary\n\n")
        content.append(f"- **Company:** {jd.company}\n")
        content.append(f"- **Role:** {jd.role}\n")
        content.append(f"- **Location:** {jd.location}\n")
        content.append(f"- **Experience Required:** {jd.experience_years} years\n")
        content.append(f"- **Key Skills:** {', '.join(jd.skills[:10])}\n\n")
        
        # Group questions by difficulty
        questions_by_difficulty = defaultdict(list)
        for question in questions:
            difficulty = question.get('difficulty', 'unknown')
            questions_by_difficulty[difficulty].append(question)
        
        # Export questions by difficulty
        for difficulty in ['easy', 'medium', 'hard']:
            if difficulty in questions_by_difficulty:
                content.append(f"## {difficulty.title()} Questions\n\n")
                
                for i, question in enumerate(questions_by_difficulty[difficulty], 1):
                    content.append(f"### {i}. {question.get('question', '')}\n\n")
                    content.append(f"**Answer:** {question.get('answer', '')}\n\n")
                    content.append(f"**Category:** {question.get('category', 'Technical')}\n")
                    content.append(f"**Skills:** {', '.join(question.get('skills', []))}\n")
                    content.append(f"**Source:** {question.get('source', 'Generated')}\n")
                    if 'relevance_score' in question:
                        content.append(f"**Relevance Score:** {question['relevance_score']:.2f}\n")
                    content.append("\n---\n\n")
        
        # Write content asynchronously
        async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
            await f.write(''.join(content))
        
        return file_path
    
    def _export_csv(self, questions: List[Dict[str, Any]], 
                   jd: JobDescription, filename_base: str) -> str:
        """
        Export questions in CSV format.
        
        Args:
            questions: List of questions
            jd: Job description object
            filename_base: Base filename
            
        Returns:
            str: Path to exported file
        """
        file_path = os.path.join(self.export_dir, f"{filename_base}.csv")
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([
                'difficulty', 'question', 'answer', 'category', 'skills', 
                'source', 'relevance_score', 'company', 'role'
            ])
            
            # Data rows
            for question in questions:
                writer.writerow([
                    question.get('difficulty', ''),
                    question.get('question', ''),
                    question.get('answer', ''),
                    question.get('category', ''),
                    ';'.join(question.get('skills', [])),
                    question.get('source', ''),
                    question.get('relevance_score', ''),
                    jd.company,
                    jd.role
                ])
        
        return file_path
    
    async def _export_csv_async(self, questions: List[Dict[str, Any]], 
                              jd: JobDescription, filename_base: str) -> str:
        """
        Export questions in CSV format asynchronously.
        
        Args:
            questions: List of questions
            jd: Job description object
            filename_base: Base filename
            
        Returns:
            str: Path to exported file
        """
        file_path = os.path.join(self.export_dir, f"{filename_base}.csv")
        
        # Prepare CSV content
        csv_content = []
        
        # Header
        csv_content.append("difficulty,question,answer,category,skills,source,relevance_score,company,role\n")
        
        # Data rows
        for question in questions:
            row = [
                question.get('difficulty', ''),
                f'"{question.get("question", "").replace('"', '""')}"',  # Escape quotes
                f'"{question.get("answer", "").replace('"', '""')}"',    # Escape quotes
                question.get('category', ''),
                f'"{";".join(question.get("skills", [])).replace('"', '""')}"',  # Escape quotes
                question.get('source', ''),
                str(question.get('relevance_score', '')),
                f'"{jd.company.replace('"', '""')}"',  # Escape quotes
                f'"{jd.role.replace('"', '""')}"'      # Escape quotes
            ]
            csv_content.append(','.join(row) + '\n')
        
        # Write content asynchronously
        async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
            await f.write(''.join(csv_content))
        
        return file_path
    
    def _export_json(self, questions: List[Dict[str, Any]], 
                    jd: JobDescription, filename_base: str) -> str:
        """
        Export questions in JSON format.
        
        Args:
            questions: List of questions
            jd: Job description object
            filename_base: Base filename
            
        Returns:
            str: Path to exported file
        """
        file_path = os.path.join(self.export_dir, f"{filename_base}.json")
        
        export_data = {
            'metadata': {
                'company': jd.company,
                'role': jd.role,
                'location': jd.location,
                'experience_years': jd.experience_years,
                'skills': jd.skills,
                'generated_at': datetime.now().isoformat(),
                'total_questions': len(questions)
            },
            'questions': questions
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        return file_path
    
    async def _export_json_async(self, questions: List[Dict[str, Any]], 
                               jd: JobDescription, filename_base: str) -> str:
        """
        Export questions in JSON format asynchronously.
        
        Args:
            questions: List of questions
            jd: Job description object
            filename_base: Base filename
            
        Returns:
            str: Path to exported file
        """
        file_path = os.path.join(self.export_dir, f"{filename_base}.json")
        
        export_data = {
            'metadata': {
                'company': jd.company,
                'role': jd.role,
                'location': jd.location,
                'experience_years': jd.experience_years,
                'skills': jd.skills,
                'generated_at': datetime.now().isoformat(),
                'total_questions': len(questions)
            },
            'questions': questions
        }
        
        # Write content asynchronously
        async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
            await f.write(json.dumps(export_data, indent=2, ensure_ascii=False))
        
        return file_path
    
    def _export_xlsx(self, questions: List[Dict[str, Any]], 
                    jd: JobDescription, filename_base: str) -> str:
        """
        Export questions in Excel format with styled headers and auto-filter.
        
        Args:
            questions: List of questions
            jd: Job description object
            filename_base: Base filename
            
        Returns:
            str: Path to exported file
        """
        file_path = os.path.join(self.export_dir, f"{filename_base}.xlsx")
        
        # Create DataFrame
        df_data = []
        for question in questions:
            df_data.append({
                'Difficulty': question.get('difficulty', ''),
                'Question': question.get('question', ''),
                'Answer': question.get('answer', ''),
                'Category': question.get('category', ''),
                'Skills': '; '.join(question.get('skills', [])),
                'Source': question.get('source', ''),
                'Relevance Score': question.get('relevance_score', ''),
                'Company': jd.company,
                'Role': jd.role
            })
        
        df = pd.DataFrame(df_data)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Interview Questions"
        
        # Add metadata sheet
        metadata_ws = wb.create_sheet("Metadata", 0)
        metadata_ws['A1'] = "Job Description Metadata"
        metadata_ws['A3'] = "Company"
        metadata_ws['B3'] = jd.company
        metadata_ws['A4'] = "Role"
        metadata_ws['B4'] = jd.role
        metadata_ws['A5'] = "Location"
        metadata_ws['B5'] = jd.location
        metadata_ws['A6'] = "Experience Required"
        metadata_ws['B6'] = f"{jd.experience_years} years"
        metadata_ws['A7'] = "Skills"
        metadata_ws['B7'] = ', '.join(jd.skills)
        metadata_ws['A8'] = "Generated At"
        metadata_ws['B8'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        metadata_ws['A9'] = "Total Questions"
        metadata_ws['B9'] = len(questions)
        
        # Style metadata headers
        for cell in metadata_ws['A1:A9']:
            cell[0].font = Font(bold=True)
        
        # Write questions data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Adjust column widths
        column_widths = {
            'A': 15,  # Difficulty
            'B': 50,  # Question
            'C': 60,  # Answer
            'D': 20,  # Category
            'E': 30,  # Skills
            'F': 15,  # Source
            'G': 15,  # Relevance Score
            'H': 25,  # Company
            'I': 25   # Role
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save workbook
        wb.save(file_path)
        
        return file_path
    
    async def _export_xlsx_async(self, questions: List[Dict[str, Any]], 
                               jd: JobDescription, filename_base: str) -> str:
        """
        Export questions in Excel format asynchronously.
        
        Args:
            questions: List of questions
            jd: Job description object
            filename_base: Base filename
            
        Returns:
            str: Path to exported file
        """
        # For Excel, we'll use the sync version since openpyxl doesn't have async support
        # but we'll run it in a thread pool to avoid blocking
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            None, 
            self._export_xlsx, 
            questions, 
            jd, 
            filename_base
        )
    
    def get_question_statistics(self) -> Dict[str, Any]:
        """
        Get statistics about the questions in the bank.
        
        Returns:
            Dict[str, Any]: Question statistics
        """
        if not self.questions:
            return {
                'total_questions': 0,
                'difficulty_distribution': {},
                'category_distribution': {},
            }
        
        stats = {
            'total_questions': len(self.questions),
            'difficulty_distribution': defaultdict(int),
            'category_distribution': defaultdict(int),
        }
        
        for question in self.questions:
            difficulty = getattr(question, 'difficulty', 'unknown')
            category = getattr(question, 'category', 'unknown')
            stats['difficulty_distribution'][difficulty] += 1
            stats['category_distribution'][category] += 1
        
        stats['difficulty_distribution'] = dict(stats['difficulty_distribution'])
        stats['category_distribution'] = dict(stats['category_distribution'])
        return stats
    
    def get_questions_by_difficulty(self) -> Dict[str, int]:
        """
        Get questions grouped by difficulty level.
        
        Returns:
            Dict[str, int]: Count of questions by difficulty
        """
        difficulty_counts = defaultdict(int)
        
        for question in self.questions:
            difficulty = getattr(question, 'difficulty', 'unknown')
            difficulty_counts[difficulty] += 1
        
        return dict(difficulty_counts)
    
    def get_average_relevance_score(self) -> float:
        """
        Get the average relevance score of all questions.
        
        Returns:
            float: Average relevance score
        """
        if not self.questions:
            return 0.0
        
        total_score = 0.0
        scored_count = 0
        
        for question in self.questions:
            if hasattr(question, 'relevance_score') and question.relevance_score is not None:
                total_score += question.relevance_score
                scored_count += 1
        
        return total_score / scored_count if scored_count > 0 else 0.0
    
    def get_export_files(self) -> List[str]:
        """
        Get list of exported files.
        
        Returns:
            List[str]: List of exported file paths
        """
        # This would typically track exported files
        # For now, return empty list
        return [] 